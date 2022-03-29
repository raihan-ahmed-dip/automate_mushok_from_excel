from turtle import title
import openpyxl
import shutil
import datetime
from win32com import client
import os

product_name = "Electronic Cook(E-Cook)-Single Burner"
unit_price = 11715.65
total_price = 11715.65
vat = 1757
price_with_vat = 13473

settings_wb = openpyxl.load_workbook('settings.xlsx')
counter_sheet = settings_wb['counter']
counter = counter_sheet['A1'].value

in_wb = openpyxl.load_workbook('input/hs_report.xlsx')
in_sheet = in_wb['hs']

max_row = in_sheet.max_row
max_col = in_sheet.max_column

#print(f"Max Row: {max_row}, Max Column: {max_col}")

row_no = 2

for i in range(1, max_row):
    # Reading data from the excel file
    name = in_sheet['A' + str(i+1)].value
    nid = in_sheet['L' + str(i+1)].value
    if nid == '(No value)':
        nid = ''
    address = in_sheet['H' + str(i+1)].value
    if address == '(No value)':
        address = ''
    product = in_sheet['E' + str(i+1)].value
    issue_date = in_sheet['O' + str(i+1)].value
    
    # Creating output files from the template
    new_file_name = "output/Mushak_6_3_" + product +"_"+ name + ".xlsx"
    shutil.copyfile('template/mushak_6_3.xlsx', new_file_name)

    date_object = datetime.datetime.strptime(issue_date, '%d/%m/%Y').date()
    month = date_object.month

    month_name = 'JAN'

    if month == 1:
        month_name = 'JAN'
    elif month == 2:
        month_name = 'FEB'
    elif month == 3:
        month_name = 'MAR'
    elif month == 4:
        month_name = 'APR'
    elif month == 5:
        month_name = 'MAY'
    elif month == 6:
        month_name = 'JUN'
    elif month == 7:
        month_name = 'JUL'
    elif month == 8:
        month_name = 'AUG'
    elif month == 9:
        month_name = 'SEP'
    elif month == 10:
        month_name = 'OCT'
    elif month == 11:
        month_name = 'NOV'
    else:
        month_name = 'DEC'

    serial = month_name + '_EC_' + str(counter)
    counter += 1

    if product == 'eCook Single':
        product_name = "Electronic Cook(E-Cook)-Single Burner"
        unit_price = 11715.65
        total_price = 11715.65
        vat = 1757
        price_with_vat = 13473
    else:
        product_name = "Electronic Cook(E-Cook)-Double Burner"
        unit_price = 19956.52
        total_price = 19956.52
        vat = 2993
        price_with_vat = 22950

    # Updating output file's data
    #invoice_id = "INV_" + str(xero_id)
    new_wb = openpyxl.load_workbook(new_file_name)
    new_sheet=new_wb.active
    new_sheet.title = name
    new_sheet['G10'] = str(name)
    #new_sheet['G11'] = datetime.date.today().strftime("%d-%b-%Y")
    new_sheet['G11'] = str(nid)
    new_sheet['G12'] = str(address)
    new_sheet['G13'] = str(address)
    new_sheet['Y10'] = str(serial)
    new_sheet['Y11'] = str(issue_date)
    new_sheet['D18'] = str(product_name)
    new_sheet['M18'] = unit_price
    new_sheet['AB18'] = price_with_vat
    new_wb.save(new_file_name)

    print(f"{i}. File Created: {new_file_name}")
    # Converting to pdf
    #in_excel_file = os.getcwd() + f"\\output\\{account_number}.xlsx"
    #out_pdf_file = os.getcwd() + f"\\output\\{account_number}.pdf"
    #print(in_excel_file)
    #excel = client.Dispatch("Excel.Application")
    #pdf_name = "output/" + str(account_number) + ".pdf"
    #sheets = excel.Workbooks.Open(in_excel_file)
    #work_sheets = sheets.Worksheets[0]
    #work_sheets.ExportAsFixedFormat(0, out_pdf_file)
    #excel.Application.Quit()
    #print(f"File Created: {out_pdf_file}")


    

